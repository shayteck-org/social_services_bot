from pyrogram import Client, filters 
from pyrogram.types import Message, ReplyKeyboardMarkup, InlineKeyboardMarkup, InlineKeyboardButton, KeyboardButton
from datetime import date, datetime, timedelta
import asyncio
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from pyrogram.errors import FloodWait, ChatAdminRequired, UserNotParticipant, RPCError
import sqlite3
from unidecode import unidecode
import matplotlib.pyplot as plt
from bidi.algorithm import get_display
import arabic_reshaper
import user_keyboard, user_message, admin_keyboard, admin_message
import openpyxl
from io import BytesIO
import logging
import requests
import json



bot = Client("MyBot", api_id=20357929, api_hash='bcdaf9715ef8d52055182e274be22ae2', bot_token="7427019552:AAH553X3PcKuYT4Mih4vGKKjmI0b2JOwurI")

conn = sqlite3.connect('botdb.db', timeout=360, check_same_thread=False, isolation_level=None)


def FirstTimeChecker(ID):
    global conn
    cur = conn.cursor()
    cur.execute(
        f'SELECT first_time FROM user WHERE user_id = "{ID}";')
    FirstTimeCheckerTemp = cur.fetchone()[0]
    cur.close()
    return FirstTimeCheckerTemp

def DataWrite(user_id, first_time):
    global conn
    cur = conn.cursor()
    cur.execute(
        f'''INSERT INTO user (user_id, first_time, balance, register, order_count) VALUES ('{user_id}', '{first_time}', '{0}', '{'❌'}', '{'0'}');''')
    conn.commit()
    cur.close()

def SetStep(ID, STEP):
    global conn
    cur = conn.cursor()
    cur.execute(
        f'''UPDATE user SET step = "{STEP}" WHERE user_id = "{ID}";''')
    conn.commit()
    cur.close()

def GetStep(ID):
    global conn
    cur = conn.cursor()
    cur.execute(f'SELECT step FROM user WHERE user_id = "{ID}";')
    GetStepTemp = cur.fetchone()[0]
    cur.close()
    return GetStepTemp

def SetLink(ID, Link):
    global conn
    cur = conn.cursor()
    cur.execute(f'UPDATE user SET link = "{Link}" WHERE user_id = "{ID}";')
    conn.commit()
    cur.close()

def GetLink(ID):
    global conn
    cur = conn.cursor()
    cur.execute(f'SELECT link FROM user WHERE user_id = "{ID}";')
    GetLinkTemp = cur.fetchone()
    cur.close()
    return GetLinkTemp[0]

def adder(id, amount):
    global conn
    cur = conn.cursor()
    cur.execute(f'SELECT balance FROM user WHERE user_id = "{id}";')
    balance = cur.fetchone()[0]
    cur.close()
    total_balance = int(balance) + int(amount)
    cur = conn.cursor()
    cur.execute(
        f'''UPDATE user SET balance = "{total_balance}" WHERE user_id = "{id}";''')
    conn.commit()
    cur.close()

def FirstTimeNo(id):
    cur = conn.cursor()
    cur.execute(
        f'''UPDATE user SET first_time = "{"no"}" WHERE user_id = "{id}";''')
    conn.commit()
    cur.close()

def admin_saver(id):
    global conn
    cur = conn.cursor()
    cur.execute(
        f'''UPDATE user SET admin = "yes" WHERE user_id = "{id}";''')
    conn.commit()
    cur.close()

def admin_remover(id):
    global conn
    cur = conn.cursor()
    cur.execute(
        f'''UPDATE user SET admin = "no" WHERE user_id = "{id}";''')
    conn.commit()
    cur.close()

def banner(id):
    global conn
    cur = conn.cursor()
    cur.execute(
        f'''UPDATE user SET ban = "yes" WHERE user_id = "{id}";''')
    conn.commit()
    cur.close()

def un_banner(id):
    global conn
    cur = conn.cursor()
    cur.execute(
        f'''UPDATE user SET ban = "no" WHERE user_id = "{id}";''')
    conn.commit()
    cur.close()

def answering(id, id2):
    global conn
    cur = conn.cursor()
    cur.execute(
        f'''UPDATE user SET answer_to = "{id2}" WHERE user_id = "{id}";''')
    conn.commit()
    cur.close()

def send_answer(id):
    global conn
    cur = conn.cursor()
    cur.execute(f'SELECT answer_to FROM user WHERE user_id = "{id}";')
    user_id = cur.fetchone()[0]
    cur.close()
    return user_id

def GetInfo():
    cur = conn.cursor()
    res = cur.execute(
                '''SELECT user_id, admin, ban, register, balance, order_count FROM user'''
            )
    result = res.fetchall()
    return result

def write_to_excel(data, filename):
        # Create a new Workbook object
    wb = openpyxl.Workbook()

    # Select the active worksheet
    ws = wb.active

    headers = ["user_id", "admin", "ban", "register", "balance", "order_count"]

    # Write headers to the first row
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data to cells, starting from the second row
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, cell_data in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell_data)
            
    # Save the workbook to a file
    wb.save(filename)
    return wb

def GetBalance(id):
    global conn
    cur = conn.cursor()
    cur.execute(f'SELECT balance FROM user WHERE user_id = "{id}";')
    balance = cur.fetchone()[0]
    cur.close()
    return balance

merchant_id = "253e4683-ae86-40f8-a627-2896973c5f24"
callback_url = "https://test.com/botredirection"

def ZarinGatewayCreate(Amount, User_ID):
    global conn
    cur = conn.cursor()
    createbody = {"merchant_id": merchant_id,
                  "amount": int(str(Amount)+"0"),
                  "description": f"{User_ID} Balance Add",
                  "callback_url": callback_url}
    result = requests.post("https://api.zarinpal.com/pg/v4/payment/request.json", json=createbody, timeout=60)
    x = result.json()["data"]["authority"]
    cur.execute(f"INSERT INTO payments (value, amount, status, user_id) VALUES ('{x}', '{Amount}', 'NotPaid', '{User_ID}');")
    conn.commit()
    cur.close()
    return x

def ZarinGatewayChecker(GatewayAdress):
    global conn
    cur = conn.cursor()
    cur.execute(f'SELECT amount FROM payments WHERE value = "{GatewayAdress}";')
    AmountTemp = cur.fetchone()[0]

    data = {
        "merchant_id": merchant_id,
        "amount": int(str(AmountTemp)+"0"),
        "authority": GatewayAdress
    }
    cur.close()
    cur = conn.cursor()
    cur.execute(
        f'SELECT status FROM payments WHERE value = "{GatewayAdress}";')
    PaymentStatus = cur.fetchone()[0]
    if PaymentStatus == "NotPaid":
        resp = requests.post(
            "https://api.zarinpal.com/pg/v4/payment/verify.json", data=data, timeout=60).json()
        try:
            statspay = resp["data"]["message"]
            ref_id = resp["data"]["ref_id"]
        except TypeError:
            return False, 0
        try:
            if statspay == "Paid":
                cur.execute(
                    f'UPDATE payments SET status = "Paid" WHERE value = "{GatewayAdress}";')
                conn.commit()
                cur.close()
                return True, ref_id
        except Exception:
            pass
    else:
        return False, "Already"


TelegramPlans = []
def TelegramPlansButtons(Plan_Type):
    global TelegramPlans
    if len(TelegramPlans) == 0:
        TelegramPlans.append([InlineKeyboardButton("🔙 بازگشت به منوی اصلی", callback_data="BackToBuyMenu")])
    TelegramPlans.insert(0, [InlineKeyboardButton((f'{Plan_Type}'), callback_data=f"PaySub!{Plan_Type}")])
    return TelegramPlans

InstagramPlans = []
def InstagramPlansButtons(Plan_Type):
    global InstagramPlans
    if len(InstagramPlans) == 0:
        InstagramPlans.append([InlineKeyboardButton("🔙 بازگشت به منوی اصلی", callback_data="BackToBuyMenu")])
    InstagramPlans.insert(0, [InlineKeyboardButton((f'{Plan_Type}'), callback_data=f"PaySub!{Plan_Type}")])
    return InstagramPlans

SpecialPlans = []
def SpecialPlansButtons(Plan_Type):
    global SpecialPlans
    if len(SpecialPlans) == 0:
        SpecialPlans.append([InlineKeyboardButton("🔙 بازگشت به منوی اصلی", callback_data="BackToBuyMenu")])
    SpecialPlans.insert(0, [InlineKeyboardButton((f'{Plan_Type}'), callback_data=f"PaySub!{Plan_Type}")])
    return SpecialPlans

MostSellPlans = []
def MostSellPlansButtons(Plan_Type):
    global MostSellPlans
    if len(MostSellPlans) == 0:
        MostSellPlans.append([InlineKeyboardButton("🔙 بازگشت به منوی اصلی", callback_data="BackToBuyMenu")])
    MostSellPlans.insert(0, [InlineKeyboardButton((f'{Plan_Type}'), callback_data=f"PaySub!{[Plan_Type]}")])
    return MostSellPlans

def ItemButtons(Plan_Type, prize, list, explain):
    list.insert(0, [InlineKeyboardButton((f'{Plan_Type}'), callback_data=f"item!{prize}!{Plan_Type}!{explain}")])
    return list

def register(id):
    global conn
    cur = conn.cursor()
    cur.execute(
        f'''UPDATE user SET register = "{"✅"}" WHERE user_id = "{id}";''')
    conn.commit()
    cur.close()


def saver_json():
    global conn, item_telegram, item_instagram, item_special, item_mostsell, TelegramPlans, InstagramPlans, SpecialPlans, MostSellPlans
    InputJSON = json.dumps({
    "item_telegram": item_telegram,
    "item_instagram": item_instagram,
    "item_special": item_special,
    "item_mostsell": item_mostsell,
    "telegram_plans": TelegramPlans,
    "instagram_plans": InstagramPlans,
    "special_plans": SpecialPlans,
    "most_sell": MostSellPlans
    })
    cur = conn.cursor()
    cur.execute('''INSERT INTO stuff (json) VALUES (?);''', (InputJSON,))
    conn.commit()
    cur.close()


def loader_json():
    global conn, item_telegram, item_instagram, item_special, item_mostsell, TelegramPlans, InstagramPlans, SpecialPlans, MostSellPlans
    cur = conn.cursor()
    cur.execute('SELECT json FROM stuff WHERE id = 1;')
    try:
        data = cur.fetchall()[0]
    except IndexError:
        return None
    
    data = json.loads(data)

    item_telegram = data[item_telegram]
    item_instagram = data[item_instagram]
    item_special = data[item_special]
    item_mostsell = data[item_mostsell]
    TelegramPlans = data[TelegramPlans]
    InstagramPlans = data[InstagramPlans]
    SpecialPlans = data[SpecialPlans]
    MostSellPlans = data[MostSellPlans]





item_telegram = []
item_instagram = []
item_special = []
item_mostsell = []
lst_tel = []
lst_ins = []
lst_special = []
lst_mostsell = []
lst_users = []
lst_admins = []
lst_ban = []








@bot.on_message(filters.private)
async def message_handler(client: Client, message: Message):
    global lst_admins, lst_users, lst_ban, lst_tel, lst_ins, lst_special, lst_mostsell, item_telegram, item_instagram, item_special, item_mostsell, lst_append
    result = GetInfo()
    for i in result:
        if i[0] != "":
            if i[0] not in lst_users:
                lst_users.append(i[0])
        if i[1] == "yes":
            if i[0] not in lst_admins:
                lst_admins.append(i[0])
        if i[2] == "yes":
            if i[0] not in lst_ban:
                lst_ban.append(i[0])
    try:
        if FirstTimeChecker(message.from_user.id) == "no":
            pass
    except Exception:
        DataWrite(message.from_user.id, first_time="yes")
        if message.text[:6] == "/start":
            if str(message.text[7:]).isdigit():
                if int(message.from_user.id) != int(message.text[7:]):
                    if GetLink(message.from_user.id) is None:
                        SetLink(message.from_user.id, int(message.text[7:]))
                        await message.reply(user_message.prize, reply_markup=user_keyboard.first_time)
                        await message.reply(user_message.home_message, reply_markup=user_keyboard.home_keyboard)
                        await bot.send_message(int(message.text[7:]), "تبریک\n\nیک نفر از طریق لینک دعوت شما وارد بات شد.\n\nازین پس شما ۵ درصد از مبلغ شارژ این کاربر را دریافت میکنید")
    try:
        await bot.get_chat_member("-1002214357298", message.from_user.id)
        if str(message.from_user.id) in lst_ban:
            await message.reply(admin_message.banned)

        elif message.text == "/start" or message.text == "بازگشت به منوی اصلی♻️":
            if str(message.from_user.id) in lst_admins:
                await message.reply(admin_message.welcome, reply_markup=admin_keyboard.home)
                SetStep(message.from_user.id, "home_admin")
            else:
                if FirstTimeChecker(message.from_user.id) == "yes":
                    await message.reply(user_message.prize, reply_markup=user_keyboard.first_time)
                    await message.reply(user_message.home_message, reply_markup=user_keyboard.home_keyboard)
                    SetStep(message.from_user.id, "home_user")
                else:
                    await message.reply(user_message.home_message, reply_markup=user_keyboard.home_keyboard)
                    SetStep(message.from_user.id, "home_user")
#all_user_part
        elif message.text == "🛍 محصولات فضـای مجازی 🛍" and GetStep(message.from_user.id) == "home_user":
            await message.reply(user_message.chose_platform, reply_markup=user_keyboard.sm_keyboard)
            SetStep(message.from_user.id, "items")

        elif GetStep(message.from_user.id) == "items" and message.text == "تلـگرام 💎":
            await message.reply(user_message.choose_service, reply_markup=InlineKeyboardMarkup(TelegramPlans))
            SetStep(message.from_user.id, "telegram")

        elif GetStep(message.from_user.id) == "items" and message.text == "اینسـتاگرام 🛒":
            await message.reply(user_message.choose_service, reply_markup=InlineKeyboardMarkup(InstagramPlans))
            SetStep(message.from_user.id, "instagram")

        elif GetStep(message.from_user.id) == "items" and message.text == "⭐️ خدمـات ویـژه ⭐️":
            await message.reply(user_message.choose_service, reply_markup=InlineKeyboardMarkup(SpecialPlans))
            SetStep(message.from_user.id, "special")

        elif GetStep(message.from_user.id) == "items" and message.text == "خدمات پر فروش":
            await message.reply(user_message.choose_service, reply_markup=InlineKeyboardMarkup(MostSellPlans))
            SetStep(message.from_user.id, "mostsell")

        elif message.text == "💡 حساب کاربری" and GetStep(message.from_user.id) == "home_user":
            for i in result:
                if str(i[0]) == str(message.from_user.id):
                    await message.reply("💡 اطلاعات کاربری شما\n\n👤 شنـــاسه کاربری :  " + str(i[0]) + "\n🛍 تعداد سفارشات :" + str(i[5]) + "\n🔥 تعداد زیر مجموعه : در دست توسعه\nوضعیت احراز :  " + str(i[3]) + "\n💎 درآمـد شما : در دست توسعه\n\n💰موجودی شما : " + str(i[4]), reply_markup=user_keyboard.user_info)

        elif message.text == "➕ افزایش موجودی" and GetStep(message.from_user.id) == "home_user":
            await message.reply(user_message.increase_balance, reply_markup=user_keyboard.increase_balance)

        elif GetStep(message.from_user.id) == "ChoosingAmount":
            if unidecode(message.text).isdigit():
                try:
                    Result = ZarinGatewayCreate(int(unidecode(message.text)), message.from_user.id)
                    AmountOfFac = f"{int(unidecode(message.text)):,}"
                    await message.reply(f"فاکتور شما به مبلغ {AmountOfFac} تومان ایجاد شد.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❇️ پرداخت فاکتور", url=f"https://www.zarinpal.com/pg/StartPay/{Result}")], [InlineKeyboardButton("🔘 بررسی پرداخت", callback_data=f"CP!{Result}!{unidecode(message.text)}")]]))
                    await message.reply("❗️ لطفا پس از پرداخت فاکتور روی دکمه (🔘 بررسی پرداخت) کلیک کنید، در غیر این صورت موجودی شما افزایش نمی یابد. 👆🏻", reply_markup=user_keyboard.home_keyboard)
                    SetStep(message.from_user.id, "home_user")
                except Exception:
                    await bot.send_message(message.from_user.id, "❌ مشکلی در درگاه پرداخت وجود دارد، لطفا دقایقی بعد دوباره امتحان کنید.")
                    SetStep(message.from_user.id, "home_user")
            else:
                await message.reply("❌ مقدار وارد شده نامعتبر میباشد، لطفا فقط عدد وارد کنید.")

        elif message.text == "🙋🏻‍♂️ درخواست مشاوره" and GetStep(message.from_user.id) == "home_user":
            await message.reply(user_message.consult)

        elif message.text == "📮 پـشتیبانـی" and GetStep(message.from_user.id) == "home_user":
            await message.reply(user_message.support, reply_markup=user_keyboard.support)

        elif message.text == "💰 شارژ رایگـان" and GetStep(message.from_user.id) == "home_user":
            await bot.send_photo(message.from_user.id, "pic.jpg", caption=f"{user_message.caption}\n\nhttp://t.me/dalal3611_bot?start={message.from_user.id}")
            await message.reply(user_message.caption2)
            #await message.reply(user_message.caption3)
        
        elif GetStep(message.from_user.id) == "send_prob" :
            await bot.copy_message("-1002214357298", from_chat_id=message.from_user.id, message_id=message.id, reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("جواب دادن", callback_data=f"supp_answer!{message.from_user.id}", url="https://t.me/dalal3611_bot")]]))
            await message.reply("پیام شما به دست پشتیبانی رسید", reply_markup=user_keyboard.home_keyboard)
            SetStep(message.from_user.id, "home_user")

        elif message.text == "🎖 درخواست نمایندگـی 🎖" and GetStep(message.from_user.id) == "home_user":
            await message.reply("در دست توسعه")

#all_admin_part
        elif message.text == "اضافه کردن ادمین جدید🧑🏻‍💻👩🏻‍💻" and GetStep(message.from_user.id) == "home_admin":
            await message.reply(admin_message.command, reply_markup=admin_keyboard.admin_set)
            SetStep(message.from_user.id, "admin_add")

        elif GetStep(message.from_user.id) == "admin_add" and message.text == "اضافه کردن ادمین جدید✅":  
            await message.reply(admin_message.user_id)
            SetStep(message.from_user.id, "admin_adding")

        elif GetStep(message.from_user.id) == "admin_adding":
            answering(message.from_user.id, message.text)
            try:
                if send_answer(message.from_user.id) in lst_admins:
                    await message.reply(admin_message.is_admin)
                elif send_answer(message.from_user.id) in lst_users:
                    admin_saver(send_answer(message.from_user.id))
                    await bot.send_message(send_answer(message.from_user.id), admin_message.you_are_admin)
                    await message.reply(admin_message.admin_set)
            except Exception:
                await message.reply(admin_message.user_not)
            SetStep(message.from_user.id, "admin_add")

        elif message.text == "حذف کردن ادمین موجود❌" and GetStep(message.from_user.id) == "admin_add":
            await message.reply(admin_message.user_id)
            SetStep(message.from_user.id, "admin_removing")

        elif GetStep(message.from_user.id) == "admin_removing":
            answering(message.from_user.id, message.text)
            try:
                if send_answer(message.from_user.id) in lst_admins:
                    await message.reply(admin_message.admin_remove)
                    admin_remover(send_answer(message.from_user.id))
                    await bot.send_message(send_answer(message.from_user.id), admin_message.you_are_not_admin)
                    lst_admins.remove(send_answer(message.from_user.id))
                    SetStep(send_answer(message.from_user.id), "home_user")
                else:
                    await message.reply(admin_message.is_not_admin)
            except Exception:
                await message.reply(admin_message.user_not)    
            SetStep(message.from_user.id, "admin_add")

        elif message.text == "مدیریت یوزرها✏️" and GetStep(message.from_user.id) == "home_admin":
            await message.reply(admin_message.user_id, reply_markup=admin_keyboard.admin_return)
            SetStep(message.from_user.id, "banning")
            
        elif GetStep(message.from_user.id) == "banning":
            answering(message.from_user.id, message.text)
            if send_answer(message.from_user.id) in lst_users:
                await message.reply(admin_message.command, reply_markup=admin_keyboard.ban_unban)        
            else:
                await message.reply(admin_message.no_user)
            SetStep(message.from_user.id, "banning")

        elif GetStep(message.from_user.id) == "answering":
            await bot.send_message(send_answer(message.from_user.id), message.text)
            await message.reply(admin_message.answer_done)

        elif message.text == "افزایش موجودی کاربر💰" and GetStep(message.from_user.id) == "home_admin":
            await message.reply(admin_message.choice_person, reply_markup=admin_keyboard.admin_return)
            SetStep(message.from_user.id, "select_person")

        elif GetStep(message.from_user.id) == "select_person":
            answering(message.from_user.id, message.text)
            if send_answer(message.from_user.id) in lst_users:
                await message.reply(admin_message.import_balance)
                SetStep(message.from_user.id, "adding")
            else:
                await message.reply(admin_message.user_not)
                SetStep(message.from_user.id, "select_person")
                await message.reply(admin_message.choice_person, reply_markup=admin_keyboard.admin_return)

        elif GetStep(message.from_user.id)[:9] == "answering":
            Data = GetStep(message.from_user.id).split("!")
            await bot.copy_message(Data[1], from_chat_id=message.from_user.id, message_id=message.id)
            SetStep(message.from_user.id, "home_admin")

        elif GetStep(message.from_user.id) == "adding":
            try:
                adder(send_answer(message.from_user.id), int(unidecode(message.text)))
                await message.reply("حساب کاربر: " + str(send_answer(message.from_user.id)) + "\nبه مقدار: " + str(message.text) + "شارژ شد.", reply_markup=admin_keyboard.home)
                await bot.send_message(send_answer(message.from_user.id), "تبریک\n\nحساب شما به مقدار: " + str(message.text) + "  شارژ شد")
                SetStep(message.from_user.id, "home_admin")
            except Exception:
                await message.reply("عبارت وارد شده درست نمی باشد\n\nلطفا فقط از اعداد بدون واحد (ریال یا تومن) استفاده کنید")
                SetStep(message.from_user.id, "adding")
                await message.reply(admin_message.import_balance)

        elif message.text == "📝اضافه کردن تاپیک جدید📝" and GetStep(message.from_user.id) == "home_admin":
            await message.reply("پلتفرم خود را انتخاب کنید", reply_markup=user_keyboard.sm_keyboard)
            SetStep(message.from_user.id, "add_but")

        elif GetStep(message.from_user.id) == "add_but" and message.text == "تلـگرام 💎":
            await message.reply("نام تاپیک را وارد کنید")
            SetStep(message.from_user.id, "item_name_tel")

        elif GetStep(message.from_user.id) == "item_name_tel":
            TelegramPlansButtons(message.text)
            item_telegram.insert(0, [])
            await message.reply("تاپیک با موفقیت ثبت شد", reply_markup=admin_keyboard.home)
            SetStep(message.from_user.id, "home_admin")

        elif GetStep(message.from_user.id) == "add_but" and message.text == "اینسـتاگرام 🛒":
            await message.reply("نام تاپیک را وارد کنید")
            SetStep(message.from_user.id, "item_name_ins")

        elif GetStep(message.from_user.id) == "item_name_ins":
            item_instagram.insert(0, [])
            InstagramPlansButtons(message.text)
            await message.reply("تاپیک با موفقیت ثبت شد", reply_markup=admin_keyboard.home)
            SetStep(message.from_user.id, "home_admin")

        elif GetStep(message.from_user.id) == "add_but" and message.text == "⭐️ خدمـات ویـژه ⭐️":
            await message.reply("نام تاپیک را وارد کنید")
            SetStep(message.from_user.id, "item_name_special")

        elif GetStep(message.from_user.id) == "item_name_special":
            SpecialPlansButtons(message.text)
            item_special.insert(0, [])
            await message.reply("تاپیک با موفقیت ثبت شد", reply_markup=admin_keyboard.home)
            SetStep(message.from_user.id, "home_admin")

        elif GetStep(message.from_user.id) == "add_but" and message.text == "خدمات پر فروش":
            await message.reply("نام تاپیک را وارد کنید")
            SetStep(message.from_user.id, "item_name_mostsell")

        elif GetStep(message.from_user.id) == "item_name_mostsell":
            MostSellPlansButtons(message.text)
            item_mostsell.insert(0, [])
            await message.reply("تاپیک با موفقیت ثبت شد", reply_markup=admin_keyboard.home)
            SetStep(message.from_user.id, "home_admin")

        elif GetStep(message.from_user.id) == "home_admin" and message.text == "✏️اضافه کردن ایتم جدید✏️":
            await message.reply("پلتفرم خود را انتخاب کنید", reply_markup=user_keyboard.sm_keyboard)
            SetStep(message.from_user.id, "add_topic")

        elif GetStep(message.from_user.id) == "add_topic" and message.text == "تلـگرام 💎":
            await message.reply("تاپیک خود را انتخاب کنید", reply_markup=InlineKeyboardMarkup(TelegramPlans))
            SetStep(message.from_user.id, "select_topic_telegram")

        elif GetStep(message.from_user.id) == "add_topic" and message.text == "اینسـتاگرام 🛒":
            await message.reply("تاپیک خود را انتخاب کنید", reply_markup=InlineKeyboardMarkup(InstagramPlans))
            SetStep(message.from_user.id, "select_topic_instagram")

        elif GetStep(message.from_user.id) == "add_topic" and message.text == "⭐️ خدمـات ویـژه ⭐️":
            await message.reply("تاپیک خود را انتخاب کنید", reply_markup=InlineKeyboardMarkup(SpecialPlans))
            SetStep(message.from_user.id, "select_topic_special")

        elif GetStep(message.from_user.id) == "add_topic" and message.text == "خدمات پر فروش":
            await message.reply("تاپیک خود را انتخاب کنید", reply_markup=InlineKeyboardMarkup(MostSellPlans))
            SetStep(message.from_user.id, "select_topic_mostsell")

        elif GetStep(message.from_user.id) == "item_name_telegram":
            await message.reply("قیمت ایتم را وارد کنید")
            lst_tel.insert(0, message.text)
            SetStep(message.from_user.id, "item_prize_telegram")

        elif GetStep(message.from_user.id) == "item_prize_telegram":
            await message.reply("توضیحات ایتم را وارد کنید")
            lst_tel.insert(1, unidecode(message.text))
            SetStep(message.from_user.id, "item_explain_telegram")

        elif GetStep(message.from_user.id) == "item_explain_telegram":
            await message.reply("ایتم شما با موفقیت ثبت شد", reply_markup=admin_keyboard.home)
            ItemButtons(lst_tel[0], lst_tel[1], item_telegram[lst_tel[2]], message.text)
            lst_tel = []
            SetStep(message.from_user.id, "home_admin")

        elif GetStep(message.from_user.id) == "item_name_instagram":
            await message.reply("قیمت ایتم را وارد کنید")
            lst_ins.insert(0, message.text)
            SetStep(message.from_user.id, "item_prize_instagram")

        elif GetStep(message.from_user.id) == "item_prize_instagram":
            await message.reply("توضیحات ایتم را وارد کنید")
            lst_ins.insert(1, unidecode(message.text))
            SetStep(message.from_user.id, "item_explain_instagram")

        elif GetStep(message.from_user.id) == "item_explain_instagram":
            await message.reply("ایتم شما با موفقیت ثبت شد", reply_markup=admin_keyboard.home)            
            ItemButtons(lst_ins[0], lst_ins[1], item_instagram[lst_ins[2]], message.text)
            lst_ins = []
            SetStep(message.from_user.id, "home_admin")

        elif GetStep(message.from_user.id) == "item_name_special":
            await message.reply("قیمت ایتم را وارد کنید")
            lst_special.insert(0, message.text)
            SetStep(message.from_user.id, "item_prize_specail")

        elif GetStep(message.from_user.id) == "item_prize_special":
            await message.reply("توضیحات ایتم را وارد کنید")
            lst_special.insert(1, unidecode(message.text))
            SetStep(message.from_user.id, "item_explain_special")

        elif GetStep(message.from_user.id) == "item_explain_special":
            await message.reply("ایتم شما با موفقیت ثبت شد", reply_markup=admin_keyboard.home)            
            ItemButtons(lst_special[0], lst_special[1], item_special[lst_special[2]], message.text)
            lst_special = []
            SetStep(message.from_user.id, "home_admin")

        elif GetStep(message.from_user.id) == "item_name_mostsell":
            await message.reply("قیمت ایتم را وارد کنید")
            lst_mostsell.insert(0, message.text)
            SetStep(message.from_user.id, "item_prize_mostsell")

        elif GetStep(message.from_user.id) == "item_prize_mostsell":
            await message.reply("توضیحات ایتم را وارد کنید")
            lst_mostsell.insert(1,unidecode(message.text))
            SetStep(message.from_user.id, "item_explain_mostsell")

        elif GetStep(message.from_user.id) == "item_explain_mostsell":
            await message.reply("ایتم شما با موفقیت ثبت شد", reply_markup=admin_keyboard.home)           
            ItemButtons(lst_mostsell[0], lst_mostsell[1], item_mostsell[lst_mostsell[2]], message.text)
            lst_mostsell = []
            SetStep(message.from_user.id, "home_admin")

        elif GetStep(message.from_user.id) == "home_admin" and message.text == "حذف تاپیک":
            await message.reply("تایپیک مدنظر خود را انتخاب کنید\n\nتوجه داشته باشید که با حذف تاپیک تمامی ایتم های موجود در تاپیک نیز حذف میگردد", reply_markup=user_keyboard.sm_keyboard)
            SetStep(message.from_user.id, "removing_topic")

        elif GetStep(message.from_user.id) == "removing_topic" and message.text == "تلـگرام 💎":
            await message.reply(user_message.choose_service, reply_markup=InlineKeyboardMarkup(TelegramPlans))
            SetStep(message.from_user.id, "telegram_remove")

        elif GetStep(message.from_user.id) == "removing_topic" and message.text == "اینسـتاگرام 🛒":
            await message.reply(user_message.choose_service, reply_markup=InlineKeyboardMarkup(InstagramPlans))
            SetStep(message.from_user.id, "instagram_remove")

        elif GetStep(message.from_user.id) == "removing_topic" and message.text == "⭐️ خدمـات ویـژه ⭐️":
            await message.reply(user_message.choose_service, reply_markup=InlineKeyboardMarkup(SpecialPlans))
            SetStep(message.from_user.id, "special_remove")

        elif GetStep(message.from_user.id) == "removing_topic" and message.text == "خدمات پر فروش":
            await message.reply(user_message.choose_service, reply_markup=InlineKeyboardMarkup(MostSellPlans))
            SetStep(message.from_user.id, "mostsell_remove")

        elif GetStep(message.from_user.id) == "home_admin" and message.text == "پیام همگانی🗣":
            await message.reply("پیام خود را وارد کنید تا به تمامی کاربران ارسال شود")
            SetStep(message.from_user.id, "send_all_message")

        elif GetStep(message.from_user.id) == "send_all_message":
            for i in lst_users:
                if i in lst_admins:
                    continue
                await bot.send_message(i, message.text)
            await message.reply("پیام شما با موفقیت ارسال شد", reply_markup=admin_keyboard.home)
            SetStep(message.from_user.id, "home_admin")

        elif GetStep(message.from_user.id) == "home_admin" and message.text == "احراز هویت کاربر":
            await message.reply(admin_message.choice_person, reply_markup=admin_keyboard.admin_return)
            SetStep(message.from_user.id, "select_person_register")

        elif GetStep(message.from_user.id) == "select_person_register":
            await message.reply(f"شما در حال احراز کاربر {message.text}می باشید", reply_markup=InlineKeyboardMarkup([InlineKeyboardButton("تایید", callback_data=f"yes!{message.text}")],[InlineKeyboardButton("🔙 بازگشت به منوی اصلی", callback_data="BackToBuyMenu")]))

        elif GetStep(message.from_user.id) == "home_admin" and message.text == "حذف ایتم":
            await message.reply("لطفا پلتفرم خود را انتخاب کنید", reply_markup=user_keyboard.sm_keyboard)
            SetStep(message.from_user.id, "removing_item")

        elif GetStep(message.from_user.id) == "removing_item":
            if message.text == "تلـگرام 💎":
                await message.reply("تاپیک خود را انتخاب کنید", reply_markup=InlineKeyboardMarkup(TelegramPlans))
                SetStep(message.from_user.id, "telegram_remove_item")
            elif message.text == "اینسـتاگرام 🛒":
                await message.reply("تاپیک خود را انتخاب کنید", reply_markup=InlineKeyboardMarkup(InstagramPlans))
                SetStep(message.from_user.id, "instagram_remove_item")
            elif message.text == "⭐️ خدمـات ویـژه ⭐️":
                await message.reply("تاپیک خود را انتخاب کنید", reply_markup=InlineKeyboardMarkup(SpecialPlans))
                SetStep(message.from_user.id, "special_remove_item")
            elif message.text == "خدمات پر فروش":
                await message.reply("تاپیک خود را انتخاب کنید", reply_markup=InlineKeyboardMarkup(MostSellPlans))
                SetStep(message.from_user.id, "mostsell_remove_item")



        elif message.text == "گزارشات📝🗂" and GetStep(message.from_user.id) == "home_admin":
            try:
                # Get user data
                data = GetInfo()
                # Write data to Excel file
                wb = write_to_excel(data, "UserTableDataBase.xlsx")
                # Get the bytes of the Excel file
                excel_bytes = BytesIO()
                wb.save(excel_bytes)
                excel_bytes.seek(0)
                # Send the Excel file as a document
                await bot.send_document(message.from_user.id, excel_bytes, file_name="UserTableDataBase.xlsx")
            except Exception as e:
                logging.error(f"Error while generating report: {e}")





    except UserNotParticipant:
        await message.reply(user_message.join_channel, reply_markup=user_keyboard.join_channel)

    except Exception:
        pass








@bot.on_callback_query()
async def callback_query(bot, CallbackQuery):
    global lst_admins, lst_users, lst_ban, lst_append
    if CallbackQuery.data == "sup":
        await CallbackQuery.edit_message_text("🔰 تیم پشتیبانی رویال عضو با افتخار آماده پاسخگویی به شما عزیزان است لطفا پیغام خود را ارسال کنید :")
        SetStep(CallbackQuery.from_user.id, "send_prob")


    elif CallbackQuery.data == "BackToMainMenu":
        if str(CallbackQuery.from_user.id) in lst_admins:
                await CallbackQuery.message.reply(admin_message.welcome, reply_markup=admin_keyboard.home)
                SetStep(CallbackQuery.from_user.id, "home_admin")
        else:
            if FirstTimeChecker(CallbackQuery.from_user.id) == "yes":
                await CallbackQuery.message.reply(user_message.prize, reply_markup=user_keyboard.first_time)
                await CallbackQuery.message.reply(user_message.home_message, reply_markup=user_keyboard.home_keyboard)
                SetStep(CallbackQuery.from_user.id, "home_user")
            else:
                await CallbackQuery.message.reply(user_message.home_message, reply_markup=user_keyboard.home_keyboard)
                SetStep(CallbackQuery.from_user.id, "home_user")

    elif CallbackQuery.data == "IncreaseGateway":
        await CallbackQuery.edit_message_text(user_message.choose_amount)
        SetStep(CallbackQuery.from_user.id, "ChoosingAmount")

    elif CallbackQuery.data[:2] == "CP":
        ReceivedData = CallbackQuery.data.split("!")
        FinalRes = ZarinGatewayChecker(ReceivedData[1])
        if FinalRes[0] is True:
            await CallbackQuery.edit_message_text(f"✅ پرداخت شما تایید شد، موجودی کیف پول شما افزایش یافت.\n\n کد پیگیری: <code>{FinalRes[1]}</code>")
            adder(CallbackQuery.from_user.id, int(ReceivedData[2]))
            try:
                adder(GetLink(CallbackQuery.from_user.id), int((int(ReceivedData[2])*0.05)))
                await bot.send_message(GetLink(CallbackQuery.from_user.id), f"💸 مبلغ {int((int(ReceivedData[2])*0.05))} تومان از طریق زیرمجموعه گیری به حساب شما افزوده شد.")
            except TypeError:
                pass
        else:
            await CallbackQuery.answer("⚠️ پرداخت شما هنوز انجام نشده است.", show_alert=True)

    elif CallbackQuery.data == "IncreaseCC":
        await CallbackQuery.edit_message_text(user_message.direct)
        
    elif CallbackQuery.data == "IncreaseCrypto":
        await CallbackQuery.edit_message_text(user_message.crypto)
        
    elif CallbackQuery.data == "check":
        try:
            await bot.get_chat_member("-1002214357298", CallbackQuery.from_user.id)
            if FirstTimeChecker(CallbackQuery.from_user.id) == "yes":
                await bot.send_message(CallbackQuery.from_user.id, user_message.prize, reply_markup=user_keyboard.first_time)
                await bot.send_message(CallbackQuery.from_user.id, user_message.home_message, reply_markup=user_keyboard.home_keyboard)
            else:
                await CallbackQuery.edit_message_text(user_message.home_message)
                await CallbackQuery.edit_message_reply_markup(user_keyboard.home_keyboard)

        except UserNotParticipant:
            await CallbackQuery.answer(user_message.not_joined)

    elif CallbackQuery.data == "increase_balance":
        await bot.send_message(CallbackQuery.from_user.id, user_message.increase_balance, reply_markup=user_keyboard.increase_balance)

    elif CallbackQuery.data == "register":
        await bot.send_message(CallbackQuery.from_user.id, user_message.register)

    elif CallbackQuery.data == "prize":
        if FirstTimeChecker(CallbackQuery.from_user.id) == "yes":
            adder(CallbackQuery.from_user.id, 500)
            FirstTimeNo(CallbackQuery.from_user.id)
            await CallbackQuery.edit_message_text(user_message.prize_claimed)
        else:
            await CallbackQuery.answer("شما جایزه خود را دریافت کردید")

    elif CallbackQuery.data[:11] == "supp_answer":
        ReceivedData = CallbackQuery.data.split("!")
        await CallbackQuery.edit_message_reply_markup(admin_keyboard.answered)
        await bot.send_message(CallbackQuery.from_user.id, admin_message.answer_message)
        SetStep(CallbackQuery.from_user.id, f"answering!{ReceivedData[1]}")

    elif CallbackQuery.data == "ban":
        if send_answer(CallbackQuery.from_user.id) not in lst_ban:
            banner(send_answer(CallbackQuery.from_user.id))
            await bot.send_message(send_answer(CallbackQuery.from_user.id), admin_message.banned)
    elif CallbackQuery.data == "unban":
        if send_answer(CallbackQuery.from_user.id) in lst_ban:
            un_banner(send_answer(CallbackQuery.from_user.id))
            lst_ban.remove(send_answer(CallbackQuery.from_user.id))
            await bot.send_message(send_answer(CallbackQuery.from_user.id), admin_message.un_banned)

    elif CallbackQuery.data[:6] == "PaySub" and GetStep(CallbackQuery.from_user.id) == "select_topic_telegram":
        Splitted = CallbackQuery.data.split("!")
        for i in range(len(TelegramPlans)):
            if str(TelegramPlans[i]) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[1]}', callback_data='PaySub!{Splitted[1]}')]"):
                for j in range(len(item_telegram)):
                    if j == i:
                        lst_tel.insert(0, j)
        await bot.send_message(CallbackQuery.from_user.id, "نام ایتم را وارد کنید")
        SetStep(CallbackQuery.from_user.id, "item_name_telegram")

    elif CallbackQuery.data[:6] == "PaySub" and GetStep(CallbackQuery.from_user.id) == "select_topic_instagram":
        Splitted = CallbackQuery.data.split("!")
        for i in range(len(InstagramPlans)):
            if str(InstagramPlans[i]) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[1]}', callback_data='PaySub!{Splitted[1]}')]"):
                for j in range(len(item_instagram)):
                    if j == i:
                        lst_ins.insert(0, j)
        await bot.send_message(CallbackQuery.from_user.id, "نام ایتم را وارد کنید")
        SetStep(CallbackQuery.from_user.id, "item_name_instagram")

    elif CallbackQuery.data[:6] == "PaySub" and GetStep(CallbackQuery.from_user.id) == "select_topic_special":
        Splitted = CallbackQuery.data.split("!")
        for i in range(len(SpecialPlans)):
            if str(SpecialPlans[i]) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[1]}', callback_data='PaySub!{Splitted[1]}')]"):
                for j in range(len(item_special)):
                    if j == i:
                        lst_special.insert(0, j)
        await bot.send_message(CallbackQuery.from_user.id, "نام ایتم را وارد کنید")
        SetStep(CallbackQuery.from_user.id, "item_name_special")

    elif CallbackQuery.data[:6] == "PaySub" and GetStep(CallbackQuery.from_user.id) == "select_topic_mostsell":
        Splitted = CallbackQuery.data.split("!")
        for i in range(len(MostSellPlans)):
            if str(MostSellPlans[i]) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[1]}', callback_data='PaySub!{Splitted[1]}')]"):
                for j in range(len(item_mostsell)):
                    if j == i:
                        lst_mostsell.insert(0, j)
        await bot.send_message(CallbackQuery.from_user.id, "نام ایتم را وارد کنید")
        SetStep(CallbackQuery.from_user.id, "item_name_mostsell")

    elif CallbackQuery.data[:6] == "PaySub" and GetStep(CallbackQuery.from_user.id) == "telegram":
        Splitted = CallbackQuery.data.split("!")
        for i in range(len(TelegramPlans)):
            if str(TelegramPlans[i]) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[1]}', callback_data='PaySub!{Splitted[1]}')]"):
                for j in range(len(item_telegram)):
                    if j == i:
                        await CallbackQuery.edit_message_reply_markup(InlineKeyboardMarkup(item_telegram[j]))

    elif CallbackQuery.data[:6] == "PaySub" and GetStep(CallbackQuery.from_user.id) == "instagram":
        Splitted = CallbackQuery.data.split("!")
        for i in range(len(InstagramPlans)):
            if str(InstagramPlans[i]) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[1]}', callback_data='PaySub!{Splitted[1]}')]"):
                for j in range(len(item_instagram)):
                    if j == i:
                        await CallbackQuery.edit_message_reply_markup(InlineKeyboardMarkup(item_instagram[j]))

    elif CallbackQuery.data[:6] == "PaySub" and GetStep(CallbackQuery.from_user.id) == "special":
        Splitted = CallbackQuery.data.split("!")
        for i in range(len(SpecialPlans)):
            if str(SpecialPlans[i]) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[1]}', callback_data='PaySub!{Splitted[1]}')]"):
                for j in range(len(item_special)):
                    if j == i:
                        await CallbackQuery.edit_message_reply_markup(InlineKeyboardMarkup(item_special[j]))

    elif CallbackQuery.data[:6] == "PaySub" and GetStep(CallbackQuery.from_user.id) == "mostsell":
        Splitted = CallbackQuery.data.split("!")
        for i in range(len(MostSellPlans)):
            if str(MostSellPlans[i]) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[1]}', callback_data='PaySub!{Splitted[1]}')]"):
                for j in range(len(item_mostsell)):
                    if j == i:
                        await CallbackQuery.edit_message_reply_markup(InlineKeyboardMarkup(item_mostsell[j]))

    elif CallbackQuery.data[:6] == "PaySub" and GetStep(CallbackQuery.from_user.id) == "telegram_remove":
        Splitted = CallbackQuery.data.split("!")
        for i in range(len(TelegramPlans)):
            if str(TelegramPlans[i]) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[1]}', callback_data='PaySub!{Splitted[1]}')]"):
                TelegramPlans.pop(i)
                item_telegram.pop(i)
                await bot.send_message(CallbackQuery.from_user.id, "تاپیک با موفقیت حذف شد", reply_markup=admin_keyboard.home)
                
    elif CallbackQuery.data[:6] == "PaySub" and GetStep(CallbackQuery.from_user.id) == "instagram_remove":
        Splitted = CallbackQuery.data.split("!")
        for i in range(len(InstagramPlans)):
            if str(InstagramPlans[i]) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[1]}', callback_data='PaySub!{Splitted[1]}')]"):
                try:  
                    InstagramPlans.pop(i)
                    item_instagram.pop(i)
                except Exception:
                    pass
                await bot.send_message(CallbackQuery.from_user.id, "تاپیک با موفقیت حذف شد", reply_markup=admin_keyboard.home)
    
    elif CallbackQuery.data[:6] == "PaySub" and GetStep(CallbackQuery.from_user.id) == "special_remove":
        Splitted = CallbackQuery.data.split("!")
        for i in range(len(SpecialPlans)):
            if str(SpecialPlans[i]) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[1]}', callback_data='PaySub!{Splitted[1]}')]"):
                SpecialPlans.pop(i)
                item_special.pop(i)
                await bot.send_message(CallbackQuery.from_user.id, "تاپیک با موفقیت حذف شد", reply_markup=admin_keyboard.home)

    elif CallbackQuery.data[:6] == "PaySub" and GetStep(CallbackQuery.from_user.id) == "mostsell_remove":
        Splitted = CallbackQuery.data.split("!")
        for i in range(len(MostSellPlans)):
            if str(MostSellPlans[i]) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[1]}', callback_data='PaySub!{Splitted[1]}')]"):
                MostSellPlans.pop(i)
                item_mostsell.pop(i)
                await bot.send_message(CallbackQuery.from_user.id, "تاپیک با موفقیت حذف شد", reply_markup=admin_keyboard.home)

    elif CallbackQuery.data[:4] == "item" and GetStep(CallbackQuery.from_user.id)[:11] != "select_item":
        splitted = CallbackQuery.data.split("!")
        await bot.send_message(CallbackQuery.from_user.id, splitted[3])
        await CallbackQuery.edit_message_text(f"نام سفارش شما :  {splitted[2]}\nهزینه سفارش شما :  {splitted[1]}\nموجودی کیف پول شما :  {GetBalance(CallbackQuery.from_user.id)}\n\nتعداد کل سفارشی که میتوانید ثبت کنید :  {int(GetBalance(CallbackQuery.from_user.id))//int(splitted[1])}")
        await CallbackQuery.edit_message_reply_markup(reply_markup=user_keyboard.put_order)
        SetStep(CallbackQuery.from_user.id, f"{splitted[1]}!{splitted[2]}")

    elif CallbackQuery.data == "one_order":
        prize = (GetStep(CallbackQuery.from_user.id).split("!"))
        if int(prize[0]) <= int(GetBalance(CallbackQuery.from_user.id)):
            adder(CallbackQuery.from_user.id, int(prize[0])*-1)
            await bot.send_message(CallbackQuery.from_user.id, "سفارش شما با موفقیت ثبت شد\n\nموجودی باقی مانده شما: " + str(GetBalance(CallbackQuery.from_user.id)), reply_markup=user_keyboard.home_keyboard)
            await bot.send_message("-1002214357298", f"کاربر {CallbackQuery.from_user.id} ایتم {prize[1]} را به قیمت {prize[0]} سفارش داد")
            SetStep(CallbackQuery.from_user.id, "home_user")
        else:
            await bot.send_message(CallbackQuery.from_user.id, "موجودی شما برای سفارش کافی نمی باشد", reply_markup=user_keyboard.home_keyboard)

    elif CallbackQuery.data == "full_order":
        if int(prize[0]) <= int(GetBalance(CallbackQuery.from_user.id)):
            prize = (GetStep(CallbackQuery.from_user.id).split("!"))
            total = int(GetBalance(CallbackQuery.from_user.id)) // int(prize[0])
            adder(CallbackQuery.from_user.id, int(int(prize[0]) * int(total) * -1))
            await bot.send_message(CallbackQuery.from_user.id, "سفارش شما با موفقیت ثبت شد\n\nموجودی باقی مانده شما: " + str(GetBalance(CallbackQuery.from_user.id)), reply_markup=user_keyboard.home_keyboard)
            await bot.send_message("-1002214357298", f"کاربر {CallbackQuery.from_user.id} ایتم {prize[1]} را به قیمت {prize[0]} سفارش داد\n\nتعداد سفارش :  {total}")
            SetStep(CallbackQuery.from_user.id, "home_user")
        else:
            await bot.send_message(CallbackQuery.from_user.id, "موجودی شما برای سفارش کافی نمی باشد", reply_markup=user_keyboard.home_keyboard)
 
    elif CallbackQuery.data[:3] == "yes":
        splitted = CallbackQuery.data.split("!")
        register(splitted[1])
        await bot.send_message(CallbackQuery.from_user.id, "کاربر با موفقیت احراز شد")
        await bot.send_message(splitted[1], "شما با موفقیت احراز شدید")
        await CallbackQuery.edit_message_reply_markup(admin_keyboard.home)
        SetStep(CallbackQuery.from_user.id, "home_admin")

    elif CallbackQuery.data == "BackToBuyMenu":
        if CallbackQuery.from_user.id in lst_admins:
            await CallbackQuery.edit_message_reply_markup(admin_keyboard.home)
            SetStep(CallbackQuery.from_user.id, "home_admin")
        elif CallbackQuery.from_user.id in lst_users:
            await CallbackQuery.edit_message_reply_markup(user_keyboard.home_keyboard)
            SetStep(CallbackQuery.from_user.id, "home_user")

    elif CallbackQuery.data[:6] == "PaySub" and GetStep(CallbackQuery.from_user.id) == "telegram_remove_item":
        Splitted = CallbackQuery.data.split("!")
        for i in range(len(TelegramPlans)):
            if str(TelegramPlans[i]) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[1]}', callback_data='PaySub!{Splitted[1]}')]"):
                for j in range(len(item_telegram)):
                    if j == i:
                        await CallbackQuery.edit_message_reply_markup(InlineKeyboardMarkup(item_telegram[j]))
                        SetStep(CallbackQuery.from_user.id, f"select_item!telegram")

    elif CallbackQuery.data[:6] == "PaySub" and GetStep(CallbackQuery.from_user.id) == "instagram_remove_item":
        Splitted = CallbackQuery.data.split("!")
        for i in range(len(InstagramPlans)):
            if str(InstagramPlans[i]) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[1]}', callback_data='PaySub!{Splitted[1]}')]"):
                for j in range(len(item_instagram)):
                    if j == i:
                        await CallbackQuery.edit_message_reply_markup(InlineKeyboardMarkup(item_instagram[j]))
                        SetStep(CallbackQuery.from_user.id, f"select_item!instagram")

    elif CallbackQuery.data[:6] == "PaySub" and GetStep(CallbackQuery.from_user.id) == "special_remove_item":
        Splitted = CallbackQuery.data.split("!")
        for i in range(len(SpecialPlans)):
            if str(SpecialPlans[i]) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[1]}', callback_data='PaySub!{Splitted[1]}')]"):
                for j in range(len(item_special)):
                    if j == i:
                        await CallbackQuery.edit_message_reply_markup(InlineKeyboardMarkup(item_special[j]))
                        SetStep(CallbackQuery.from_user.id, f"select_item!special")

    elif CallbackQuery.data[:6] == "PaySub" and GetStep(CallbackQuery.from_user.id) == "mostsell_remove_item":
        Splitted = CallbackQuery.data.split("!")
        for i in range(len(MostSellPlans)):
            if str(MostSellPlans[i]) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[1]}', callback_data='PaySub!{Splitted[1]}')]"):
                for j in range(len(item_mostsell)):
                    if j == i:
                        await CallbackQuery.edit_message_reply_markup(InlineKeyboardMarkup(item_mostsell[j]))
                        SetStep(CallbackQuery.from_user.id, f"select_item!mostsell")

    elif CallbackQuery.data[:4] == "item" and GetStep(CallbackQuery.from_user.id)[:11] == "select_item":
        Splitted = CallbackQuery.data.split("!")
        Splitted2 = GetStep(CallbackQuery.from_user.id).split("!")
        if str(Splitted2[1]) == "telegram":
            try:
                for i in range(len(item_telegram)):
                    for j in item_telegram[i]:
                        if str(j) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[2]}', callback_data='item!{Splitted[1]}!{Splitted[2]}!{Splitted[3]}')]"):
                            (item_telegram[i]).remove(j)
                            await bot.send_message(CallbackQuery.from_user.id, "ایتم با موفقیت حذف شد", reply_markup=admin_keyboard.home)
                            SetStep(CallbackQuery.from_user.id, "home_admin")
            except Exception:
                pass

        elif str(Splitted2[1]) == "instagram":
            try:
                for i in range(len(item_instagram)):
                    for j in item_instagram[i]:
                        if str(j) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[2]}', callback_data='item!{Splitted[1]}!{Splitted[2]}!{Splitted[3]}')]"):
                            (item_instagram[i]).remove(j)
                            await bot.send_message(CallbackQuery.from_user.id, "ایتم با موفقیت حذف شد", reply_markup=admin_keyboard.home)
                            SetStep(CallbackQuery.from_user.id, "home_admin")
            except Exception:
                pass

        elif str(Splitted2[1]) == "special":
            try:
                for i in range(len(item_special)):
                    for j in item_special[i]:
                        if str(j) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[2]}', callback_data='item!{Splitted[1]}!{Splitted[2]}!{Splitted[3]}')]"):
                            (item_special[i]).remove(j)
                            await bot.send_message(CallbackQuery.from_user.id, "ایتم با موفقیت حذف شد", reply_markup=admin_keyboard.home)
                            SetStep(CallbackQuery.from_user.id, "home_admin")
            except Exception:
                pass


        elif str(Splitted2[1]) == "mostsell":
            try:
                for i in range(len(item_mostsell)):
                    for j in item_mostsell[i]:
                        if str(j) == str(f"[pyrogram.types.InlineKeyboardButton(text='{Splitted[2]}', callback_data='item!{Splitted[1]}!{Splitted[2]}!{Splitted[3]}')]"):
                            (item_mostsell[i]).remove(j)
                            await bot.send_message(CallbackQuery.from_user.id, "ایتم با موفقیت حذف شد", reply_markup=admin_keyboard.home)
                            SetStep(CallbackQuery.from_user.id, "home_admin")
            except Exception:
                pass

print("Bot Is On")
loader_json()
bot.run()
saver_json()
print("Bot Is Off")